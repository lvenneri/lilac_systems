"""Parse an experiment config Excel workbook into Python dicts."""

import openpyxl


def _parse_bool(val):
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().lower() in ("yes", "true", "1")
    return bool(val)


def _to_number(val):
    """Try to convert to float; return original if not numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    try:
        return float(val)
    except (ValueError, TypeError):
        return val


def _to_float(val, default=0):
    """Convert to float, returning *default* if not numeric."""
    n = _to_number(val)
    return n if isinstance(n, (int, float)) else default


def _sheet_rows(wb, sheet_name):
    """Yield (header_list, data_rows) for a sheet. Row 1 = description (skip),
    Row 2 = headers, Row 3+ = data."""
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        return [], []
    headers = [str(h).strip() if h else "" for h in rows[1]]
    data = rows[2:]
    return headers, data


def _row_to_dict(headers, row):
    """Convert a row tuple to a dict keyed by headers."""
    return {h: row[i] if i < len(row) else None for i, h in enumerate(headers) if h}


def load_config(filepath):
    """Parse the Excel config and return a structured dict.

    Returns dict with keys: instruments, channels, control_loops,
    interlocks, logging, settings.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    config = {}

    # --- Instruments ---
    headers, data = _sheet_rows(wb, "Instruments")
    instruments = {}
    for row in data:
        d = _row_to_dict(headers, row)
        name = d.get("Instrument Name")
        if not name:
            continue
        name = str(name).strip()
        instruments[name] = {
            "type": str(d.get("Type", "simulated")).strip().lower(),
            "address": str(d.get("Address / Device", "")) if d.get("Address / Device") else "",
            "query_command": str(d.get("Query Command", "")) if d.get("Query Command") else "",
            "poll_rate": _to_float(d.get("Poll Rate (s)"), 0.1),
            "timeout": _to_float(d.get("Timeout (s)"), 5),
            "enabled": _parse_bool(d.get("Enabled", True)),
            "notes": str(d.get("Notes", "")) if d.get("Notes") else "",
        }
    config["instruments"] = instruments

    # --- Channels ---
    headers, data = _sheet_rows(wb, "Channels")
    channels = {}
    for row in data:
        d = _row_to_dict(headers, row)
        name = d.get("Channel Name")
        if not name:
            continue
        name = str(name).strip()
        channels[name] = {
            "instrument": str(d.get("Instrument", "")).strip(),
            "channel_id": str(d.get("Channel ID", "")).strip(),
            "direction": str(d.get("Direction", "input")).strip().lower(),
            "signal_type": str(d.get("Signal Type", "")).strip().lower(),
            "units": str(d.get("Units", "")) if d.get("Units") else "",
            "slope": _to_float(d.get("Slope"), 1),
            "offset": _to_float(d.get("Offset"), 0),
            "min": _to_number(d.get("Min Value")) if d.get("Min Value") is not None else 0,
            "max": _to_number(d.get("Max Value")) if d.get("Max Value") is not None else 100,
            "enabled": _parse_bool(d.get("Enabled", True)),
            "control_options": str(d.get("Control Options", "")).strip() if d.get("Control Options") else "",
            "notes": str(d.get("Notes", "")) if d.get("Notes") else "",
        }
    config["channels"] = channels

    # --- Build channel_options from Signal Type for NI cDAQ drivers ---
    # Signal Type values: "tc_K", "tc_T", "rtd_PT_3851", "rtd_PT_3851_4wire",
    #                     "poly_c0,c1,c2,...", etc.
    for ch_name, ch_cfg in channels.items():
        sig = ch_cfg.get("signal_type", "")
        if not sig:
            continue
        inst_name = ch_cfg.get("instrument", "")
        channel_id = ch_cfg.get("channel_id", "")
        if not inst_name or not channel_id or inst_name not in instruments:
            continue
        opts = {}
        sig_upper = sig.upper().replace(" ", "_")
        if sig_upper.startswith("TC_"):
            # e.g. "tc_K", "tc_T", "tc_J"
            opts["tc_type"] = sig_upper[3:]
        elif sig_upper.startswith("RTD_"):
            # e.g. "rtd_PT_3851", "rtd_PT_3851_4wire", "rtd_PT_3750_2wire"
            parts = sig_upper[4:]  # "PT_3851_4WIRE"
            # Check for trailing wire config
            for wire in ("_2WIRE", "_3WIRE", "_4WIRE"):
                if parts.endswith(wire):
                    opts["rtd_wires"] = wire[-5]  # "2", "3", or "4"
                    parts = parts[: -len(wire)]
                    break
            if parts:
                opts["rtd_type"] = parts
        elif sig.lower().startswith("poly_"):
            # e.g. "poly_0.5,1.2,0.003" -> coefficients [0.5, 1.2, 0.003]
            coeff_str = sig[5:]  # strip "poly_"
            try:
                coeffs = [float(c.strip()) for c in coeff_str.split(",")]
                opts["coefficients"] = coeffs
            except ValueError:
                pass
        if opts:
            inst = instruments[inst_name]
            if "channel_options" not in inst:
                inst["channel_options"] = {}
            inst["channel_options"][channel_id] = {
                **inst.get("channel_options", {}).get(channel_id, {}),
                **opts,
            }

    # --- Control Loops ---
    headers, data = _sheet_rows(wb, "Control Loops")
    control_loops = {}
    for row in data:
        d = _row_to_dict(headers, row)
        name = d.get("Loop Name")
        if not name:
            continue
        name = str(name).strip()
        control_loops[name] = {
            "pv_channel": str(d.get("Process Variable", "")).strip(),
            "setpoint": _to_number(d.get("Setpoint")) or 0,
            "sp_units": str(d.get("SP Units", "")) if d.get("SP Units") else "",
            "output_channel": str(d.get("Output Channel", "")).strip(),
            "out_min": _to_number(d.get("Out Min")) if d.get("Out Min") is not None else 0,
            "out_max": _to_number(d.get("Out Max")) if d.get("Out Max") is not None else 100,
            "kp": _to_number(d.get("Kp")) or 0,
            "ki": _to_number(d.get("Ki")) or 0,
            "kd": _to_number(d.get("Kd")) or 0,
            "sample_time": _to_number(d.get("Sample Time (s)")) or 0.1,
            "mode": str(d.get("Mode", "manual")).strip().lower(),
            "enabled": _parse_bool(d.get("Enabled", True)),
            "notes": str(d.get("Notes", "")) if d.get("Notes") else "",
        }
    config["control_loops"] = control_loops

    # --- Interlocks ---
    headers, data = _sheet_rows(wb, "Interlocks")
    interlocks = []
    for row in data:
        d = _row_to_dict(headers, row)
        name = d.get("Interlock Name")
        if not name:
            continue
        # Parse compound actions (semicolon-separated)
        raw_actions = str(d.get("Action", "alarm")).strip().lower()
        raw_targets = str(d.get("Target", "")).strip() if d.get("Target") else ""
        raw_values = str(d.get("Action Value", "")).strip() if d.get("Action Value") is not None else ""
        action_list = [a.strip() for a in raw_actions.split(";") if a.strip()]
        target_list = [t.strip() for t in raw_targets.split(";")]
        value_list = [v.strip() for v in str(raw_values).split(";")]
        # Pad shorter lists to match action_list length
        while len(target_list) < len(action_list):
            target_list.append("")
        while len(value_list) < len(action_list):
            value_list.append("")
        actions = []
        for i, act in enumerate(action_list):
            val = _to_number(value_list[i]) if value_list[i] else None
            actions.append({"action": act, "target": target_list[i], "value": val})

        # Parse recovery actions (same format)
        raw_rec_actions = str(d.get("Recovery Action", "")).strip().lower() if d.get("Recovery Action") else ""
        raw_rec_targets = str(d.get("Recovery Target", "")).strip() if d.get("Recovery Target") else ""
        raw_rec_values = str(d.get("Recovery Value", "")).strip() if d.get("Recovery Value") is not None else ""
        recovery = []
        if raw_rec_actions:
            rec_act_list = [a.strip() for a in raw_rec_actions.split(";") if a.strip()]
            rec_tgt_list = [t.strip() for t in raw_rec_targets.split(";")]
            rec_val_list = [v.strip() for v in str(raw_rec_values).split(";")]
            while len(rec_tgt_list) < len(rec_act_list):
                rec_tgt_list.append("")
            while len(rec_val_list) < len(rec_act_list):
                rec_val_list.append("")
            for i, act in enumerate(rec_act_list):
                val = _to_number(rec_val_list[i]) if rec_val_list[i] else None
                recovery.append({"action": act, "target": rec_tgt_list[i], "value": val})

        interlocks.append({
            "name": str(name).strip(),
            "channel": str(d.get("Channel", "")).strip(),
            "condition": str(d.get("Condition", ">")).strip(),
            "threshold": _to_number(d.get("Threshold")) or 0,
            "actions": actions,
            # Keep legacy single-action fields for backwards compat
            "action": action_list[0] if action_list else "alarm",
            "target": target_list[0] if target_list else "",
            "action_value": _to_number(value_list[0]) if value_list and value_list[0] else None,
            "recovery": recovery,
            "latch": _parse_bool(d.get("Latch", False)),
            "enabled": _parse_bool(d.get("Enabled", True)),
            "group": str(d.get("Group", "")).strip() if d.get("Group") else "",
            "notes": str(d.get("Notes", "")) if d.get("Notes") else "",
        })
    config["interlocks"] = interlocks

    # --- Logging ---
    headers, data = _sheet_rows(wb, "Logging")
    logging_cfg = {}
    for row in data:
        d = _row_to_dict(headers, row)
        ch = d.get("Channel")
        if not ch:
            continue
        ch = str(ch).strip()
        logging_cfg[ch] = {
            "log_csv": _parse_bool(d.get("Log to CSV", True)),
            "display": _parse_bool(d.get("Display on UI", True)),
            "display_format": str(d.get("Display Format", "")) if d.get("Display Format") else "",
            "decimal_places": _to_number(d.get("Decimal Places")),
            "alarm_low": _to_number(d.get("Alarm Low")),
            "alarm_high": _to_number(d.get("Alarm High")),
            "notes": str(d.get("Notes", "")) if d.get("Notes") else "",
        }
    config["logging"] = logging_cfg

    # --- Settings ---
    headers, data = _sheet_rows(wb, "Settings")
    settings = {}
    for row in data:
        d = _row_to_dict(headers, row)
        key = d.get("Setting")
        val = d.get("Value")
        if not key or val is None or str(val).strip() == "":
            continue
        key_str = str(key).strip()
        # Skip section headers (emoji rows)
        if key_str and any(ord(c) > 0x2000 for c in key_str):
            continue
        num = _to_number(val)
        settings[key_str] = num if isinstance(num, (int, float)) else str(val).strip()
    config["settings"] = settings

    # --- Step Series ---
    headers, data = _sheet_rows(wb, "Step Series")
    step_series = []
    step_columns = []
    if headers and data:
        # Check for tolerance row (first data row with "Tolerance" in Step # column)
        tolerance_row = None
        step_data = data
        if data:
            first = _row_to_dict(headers, data[0])
            first_step = str(first.get("Step #", "")).strip().lower()
            if first_step == "tolerance":
                tolerance_row = first
                step_data = data[1:]

        for h in headers:
            h_stripped = str(h).strip() if h else ""
            if not h_stripped or h_stripped in ("Step #", "Hold Time (s)", "Name"):
                continue
            col_info = {"header": h_stripped}
            if h_stripped.startswith("SP: "):
                remainder = h_stripped[4:].strip()
                if "(" in remainder:
                    pv_name = remainder[:remainder.index("(")].strip()
                else:
                    pv_name = remainder
                col_info["type"] = "pid_setpoint"
                col_info["pv_channel"] = pv_name
            elif h_stripped.startswith("Watch: "):
                remainder = h_stripped[7:].strip()
                if "(" in remainder:
                    ch_name = remainder[:remainder.index("(")].strip()
                else:
                    ch_name = remainder
                col_info["type"] = "watch"
                col_info["channel_name"] = ch_name
            else:
                col_info["type"] = "output_channel"
                if "(" in h_stripped:
                    col_info["channel_name"] = h_stripped[:h_stripped.index("(")].strip()
                else:
                    col_info["channel_name"] = h_stripped
            # Extract tolerance for this column
            tol = None
            if tolerance_row:
                tol = _to_number(tolerance_row.get(h_stripped))
            col_info["tolerance"] = tol
            step_columns.append(col_info)

        for row in step_data:
            d = _row_to_dict(headers, row)
            step_num = d.get("Step #")
            if step_num is None:
                continue
            hold_time = _to_number(d.get("Hold Time (s)")) or 0
            setpoints = {}
            for col in step_columns:
                raw_val = d.get(col["header"])
                if isinstance(raw_val, str) and raw_val.strip().upper() in ("NA", "N/A", ""):
                    continue
                val = _to_number(raw_val)
                if val is not None and isinstance(val, (int, float)):
                    if col["type"] == "pid_setpoint":
                        setpoints[col["header"]] = {
                            "type": "pid_setpoint",
                            "pv_channel": col["pv_channel"],
                            "value": val,
                        }
                    elif col["type"] == "watch":
                        setpoints[col["header"]] = {
                            "type": "watch",
                            "channel_name": col["channel_name"],
                            "value": val,
                        }
                    else:
                        setpoints[col["header"]] = {
                            "type": "output_channel",
                            "channel_name": col["channel_name"],
                            "value": val,
                        }
            name = str(d.get("Name", "")).strip()
            step_series.append({
                "step_num": int(step_num),
                "name": name,
                "hold_time": float(hold_time),
                "setpoints": setpoints,
            })
    config["step_series"] = step_series
    config["step_columns"] = step_columns

    wb.close()
    return config
